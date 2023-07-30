import os
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from PIL import Image

from analytics import Univariate, Bivariate, CombinedDataset
from utils import FileExt

path = os.path.realpath(__file__)
outputdir = os.path.abspath(f"{path}/../_output/")

def save_dataset(dataset: Univariate | Bivariate | CombinedDataset, with_ext = FileExt.txt):
    if with_ext == FileExt.txt:
        filepath = f"{outputdir}/output.txt"
        with open(filepath, "w") as file:
            file.write(str(dataset))
    else:
        print(f"Print dataset: extension not supported ({with_ext})")

def save_dataset_xlsx(dataset: CombinedDataset):
    filepath = f"{outputdir}/output.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for col in range(ord('A'), ord('Z')+1):
        col_letter = chr(col)
        sheet.column_dimensions[col_letter].width = 12

    for row in range(1, 41):
        sheet.row_dimensions[row].height = 32

    header_fill = PatternFill(start_color='bfd6ba', end_color='bfd6ba', fill_type='solid')
    header_font = Font(size=12, bold=True)
    hc_names = ["A1:D1", "F1:L1"]
    hc_text = ["N", "X", "X ранж", "Y", "N", "GI", "Ch", "NCh", "Cha", "NCha", "Means"]
    header_cells = [cell for s in hc_names for rows in sheet[s] for cell in rows]

    for i, cell in enumerate(header_cells):
        cell.value = hc_text[i]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    block_one = dataset.value_block_one
    for i, cell in enumerate([c for row in sheet[f"A2:D{int(len(block_one) / 4)+1}"] for c in row]):
        cell.value = block_one[i]

    block_two = dataset.value_block_two
    for i, cell in enumerate([c for row in sheet[f"F2:L{int(len(block_two) / 7)+1}"] for c in row]):
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.value = block_two[i]

    workbook.save(filepath)

def writetxt(s, name = "default", with_ext = FileExt.txt):
    if with_ext == FileExt.txt:
        filepath = f"{outputdir}/{name}.txt"
        with open(filepath, "w") as file:
            file.write(str(s))
    else:
        print(f"Print dataset: extension not supported ({with_ext})")

