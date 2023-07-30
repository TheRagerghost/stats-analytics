import os
import openpyxl
from utils import round_full

path = os.path.realpath(__file__)
inputdir = os.path.abspath(f"{path}/../_input/")

def load_xlsx(filename: str):
    filepath = f"{inputdir}/{filename}.xlsx"
    if os.path.isfile(filepath):
        print(f'File "{filepath}" is loading.')
    else:
        print(f'The file "{filepath}" does not exist.')
        return
    workbook = openpyxl.load_workbook(filepath)
    return get_values(workbook)

def load_default_xlsx():
    filepath = f"{inputdir}/default.xlsx"
    if os.path.isfile(filepath):
        print(f'File "{filepath}" is loading.')
    else:
        print(f'The file "{filepath}" does not exist.')
        return
    workbook = openpyxl.load_workbook(filepath)
    return get_values(workbook)

def get_values(workbook):
    sheet = workbook.active
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'X':
                x_row = cell.row
                x_col = cell.column
            elif cell.value == 'Y':
                y_row = cell.row
                y_col = cell.column
    x_values = []
    y_values = []
    for row in sheet.iter_rows(min_row=x_row+1, min_col=x_col, values_only=True):
        if isinstance(row[0], (int, float)):
            x_values.append(round_full(float(row[0]), 4))
    for row in sheet.iter_rows(min_row=y_row+1, min_col=y_col, values_only=True):
        if isinstance(row[0], (int, float)):
            y_values.append(round_full(float(row[0]), 4))
#    print(f"\nX values: {x_values}\nY values: {y_values}")
    return x_values, y_values